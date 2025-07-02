from openpyxl.utils import coordinate_to_tuple
import os
from datetime import datetime

def validate_cell_value(cell, value_type):
    if cell.value is None:
        return None, True
    
    if value_type == 'v-number':
        if isinstance(cell.value, int):
            return cell.value, True
    
        try:
            return int(cell.value), True
        except:
            pass
    
    if value_type == 'v-string':    
        return cell.value, True
    
    return cell.value, False

def copy_cell_value(master_ws, master_row, master_col, input_ws, input_row, input_col, value_type, logger, input_file):
    input_ws_cell = input_ws.cell(input_row, input_col)
    input_ws_cell_value, is_valid = validate_cell_value(input_ws_cell, value_type)
    if is_valid:
        master_ws.cell(master_row, master_col).value = input_ws_cell_value
        
    else:
        logger.error(f'{input_file} | Wrong value of cell: {input_ws.cell(input_row, input_col).coordinate} | [{input_ws_cell.value}]')

def get_keys_in_worksheet(ws):
    keys = {}

    # get number of rows
    num_rows = ws.max_row

    # loop through each row
    for row in range(1, num_rows + 1):
        if ws.cell(row, 1).value:
            if not ws.cell(row,2).font.bold:
                keys[ws.cell(row, 1).value] = {
                                                'row': row,
                                                'row_type': 'detail'
                                                }
            else:
                keys[ws.cell(row, 1).value] = {
                                                'row': row,
                                                'row_type': 'summary'
                                                }
    try:
        keys.pop('TOTAL DAV')
    except:
        pass

    return keys
def get_input_files(sharepoint_folder, data_period = None, logger = None):
    input_files = []
    if not data_period:
        data_period = datetime.now().strftime('%Y%m')

    sub_folders_branch_group = os.listdir(sharepoint_folder)
    try:
        sub_folders_branch_group.remove('1. DAV')
    except:
        pass

    for branch_group in sub_folders_branch_group:
        branch_folders = os.listdir(os.path.join(sharepoint_folder, branch_group))

        for branch_folder in branch_folders:
            # check foler exist
            month_folder = os.path.join(sharepoint_folder,branch_group,branch_folder, data_period)
            if os.path.exists(month_folder):
                file_list = os.listdir(month_folder)
                file_list = [file for file in file_list if file.endswith('.xlsx') and 'FCT' in file]
                input_files.append(os.path.join(month_folder, file_list[0]))
                if len(file_list) != 1:
                    logger.error(f'ERROR ------ {month_folder} has {len(file_list)} file(s)')
                
            else:
                logger.error(f'{month_folder} does not exist')
    return input_files

def get_master_file(sharepoint_folder, data_period = None):
    if not data_period:
        data_period = datetime.now().strftime('%Y%m')
    sharepoint_folder = os.path.abspath(sharepoint_folder)
    master_folder = os.path.join(sharepoint_folder, '1. DAV')
    master_folder = os.path.join(master_folder, data_period)

    master_files = os.listdir(master_folder)
    for file in master_files:
        file_upper = file.upper()
        if file_upper.endswith('.XLSX') and 'FCT' in file_upper and 'MASTER' in file_upper:
            return os.path.join(master_folder, file)
    return None